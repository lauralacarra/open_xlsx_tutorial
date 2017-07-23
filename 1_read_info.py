from openpyxl import load_workbook
wb = load_workbook('sosz17info.xlsx')
sheets = wb.get_sheet_names()
print(sheets)
print("For each sheet")
for sheet in sheets:
    ws = wb[sheet]
    print(ws.title)
    print("MAX column " + str(ws.max_column))
    print("MAX row " + str(ws.max_row))
    iterrows = iter(ws.iter_rows(min_row=1,
                                 max_col=ws.max_column,
                                 max_row=ws.max_row
                                 )
                    )
    # skip first line
    # next(iterrows)
    for row in iterrows:
        row_string = ""
        for cell in row:
            #  print("Columna")
            row_string = row_string + " | " + str(cell.value)
        print(row_string)
